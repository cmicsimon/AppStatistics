import pymysql
import pandas as pd
import os
from openpyxl import load_workbook

#查询类型与数据库对应表名
tables = {"amount":"m_amount", "order":"m_order", "user":"m_user",
            "province":"m_province", "nationwide":"m_nationwide",
            "info":"application"}
#查询类型与excel表对应sheet名
sheets = {"amount":"应用流水金额","order":"应用流水订单量","user":"应用流水人数",
          "province":"分省分应用流水","nationwide":"全国","info":"计费类型"}
#查询“金额”“订单量”“人数”时对应的字段
field_normal = """,date,time0,time1,time2,time3,time4,time5,time6,time7,time8,time9,time10,
                time11,time12,time13,time14,time15,time16,time17,time18,time19,
                time20,time21,time22,time23"""
#查询“省份”时对应的字段
field_province = ",province,date,pre_user,pre_order,pre_amount,cur_user,cur_order,cur_amount"
#查询“全国”时对应的字段
field_nationwide = ",date,pre_user,pre_order,pre_amount,cur_user,cur_order,cur_amount"
# “日期”条件
where_date = " and {table}.date >= {start} and {table}.date <= {end}"
#“省份”条件
where_province = " and province in {province}"

#输入应用id
def getID():
    id_invalid = True
    while id_invalid:
        id = input("请输入应用id(id0,id1,...)：")
        if "," in id:
            items = id.split(",")
            for item in items:
                if item.isdigit():
                    if item == items[-1]:
                        id_invalid = False
                else:
                    print("应用id错误")
                    break
        elif id.isdigit():
            id_invalid = False
        else:
            print("应用id错误")
    return id

#输入日期
def getDate():
    while True:
        date = input("请输入日期(yyyymmdd-yyyymmdd)：")
        if "-" in date:
            date = date.split("-")
            start = date[0]
            end = date[1]
            if start.isdigit() and end.isdigit():
                break;
            else:
                print("日期格式错误")
        elif date.isdigit():
            start = date
            end = date
            break;
        else:
            print("日期格式错误")
    date = (start, end)
    return date

#输入查询类型
def getType():
    while True:
        type = input("请输入查询类型(amount/order/user/province/nationwide/info)：")
        if type in tables.keys():
            return type
        else:
            print("查询类型错误")

#输入查询省份
def getProvince():
    provinces = input("请输入省份(province0,province1,...)：")
    if "," in provinces:
        provinces = provinces.split(",")
        print(provinces)
        provinces = tuple(provinces)
    else:
        provinces = "(\'" + provinces + '\')'
    return provinces

#输出excel文件
def output(data, type):
    while True:
        file = input("请输入导出文件名:")
        if file.endswith(".xlsx"):
            break;
        else:
            print("文件格式错误")
    writer = pd.ExcelWriter(file)
    if type in ("amount","order","user"):
        list = []
        cols = ["日期","应用名称","应用id","AP代码","AP名称"]
        for i in range(24):
            cols.append(str(i))
        for row in data:
            col = {}
            col[cols[0]] = str(row[5])
            col[cols[1]] = row[1]
            col[cols[2]] = row[0]
            col[cols[3]] = row[2]
            col[cols[4]] = row[3]
            for i in range(24):
                col[str(i)] = row[i+6]
            list.append(col)
    elif type == "province":
        list = []
        cols = ["日期","省份","应用名称","应用id","AP代码","AP名称",
                "前一日订购人数","前一日订单数","前一日金额",
                "当日订购人数","当日订单数","当日金额"]
        for row in data:
            col = {}
            col[cols[0]] = str(row[6])
            col[cols[1]] = row[5]
            col[cols[2]] = row[1]
            col[cols[3]] = row[0]
            col[cols[4]] = row[2]
            col[cols[5]] = row[3]
            for i in range(6):
                col[cols[i+6]] = row[i+7]
            list.append(col)
    elif type == "nationwide":
        list = []
        cols = ["日期","应用名称","应用id","AP代码","AP名称",
                "前一日订购人数","前一日订单数","前一日金额",
                "当日订购人数","当日订单数","当日金额"]
        for row in data:
            col = {}
            col[cols[0]] = str(row[5])
            col[cols[1]] = row[1]
            col[cols[2]] = row[0]
            col[cols[3]] = row[2]
            col[cols[4]] = row[3]
            for i in range(6):
                col[cols[i+5]] = row[i+6]
            list.append(col)
    elif type=="info":
        list = []
        cols = ["应用id","应用名称","AP代码","AP名称","计费类型"]
        for row in data:
            col = {}
            for i in range(5):
                col[cols[i]] = row[i]
            list.append(col)
    df = pd.DataFrame(data=list)
    df = df.ix[:, cols]
    if os.path.exists(file) != True:
        df.to_excel(excel_writer=writer, sheet_name=sheets[type], index=None)
    else:
        book = load_workbook(writer.path)
        writer.book = book
        sheet_names = writer.book.sheetnames
        if sheets[type] in sheet_names:#判断sheet是否存在，存在则先删除
            del writer.book[sheets[type]]
        df.to_excel(excel_writer=writer, sheet_name=sheets[type], index=None)
    writer.save()

#查询数据
def query():
    db = pymysql.connect(host="localhost", user="test", passwd="test123",
                         db="cmic_application_statistics", port=3306, charset="utf8")
    cursor = db.cursor()
    id = getID()#获取应用id
    type = getType()#获取查询类型
    table = tables[type]#获取对应表名
    basic_table = "application,"
    if type == "info":
        where = ""
        field = ""
        basic_table = ""
    else:
        date = getDate()
        if type == "province":
            province = getProvince()
            field = field_province
            global where_date
            where_date = where_date + where_province.format(province=province)
        elif type == "nationwide":
            field = field_nationwide
        else:
            field = field_normal
        where = where_date.format(table=table, start=date[0], end=date[1])

    sql = """select application.*{field} from {basic_table}{table}
    where application.app_id in ({id}) and {table}.app_id = application.app_id{where};""" \
        .format(basic_table=basic_table, table=table, field=field, id=id, where=where)
    print(sql)
    try:
        cursor.execute(sql)
        result = cursor.fetchall()
        print(result)
    except:
        print("error")
        db.rollback()
    if(len(result) > 0):
        output(result, type)
    else:
        print("查无结果")
    cursor.close()
    db.close()

query()


